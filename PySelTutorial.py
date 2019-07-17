"""
from selenium import webdriver
import os

class RunChromeTests():
    # http://chromedriver.storage.googleapis.com/index.html

    def test(self):
        driverLocation = "D:/automation/chromedriver"
        os.environ["webdriver.chrome.driver"] = driverLocation
        # Instantiate Chrome Browser Command
        driver = webdriver.Chrome(driverLocation)
        # Open the provided URL
        driver.get("http://www.letskodeit.com")

ff = RunChromeTests()
ff.test()
"""

#XPATH-Relative
"""
package com.edureka.xpath;

import org.openqa.selenium.By;
import org.openqa.selenium.WebDriver;
import org.openqa.selenium.WebElement;
import org.openqa.selenium.chrome.ChromeDriver;

public class LocatorByRelativeXpath {

	public static WebDriver driver = null;

	public static void main(String[] args) 
	{
		System.setProperty("webdriver.chrome.driver", ".\\Drivers\\chromedriver.exe");
		driver = new ChromeDriver();
		driver.manage().window().maximize();
		driver.get("https://www.facebook.com/");
		
		String relativeXpath = "//tagname[@attribute = 'value'] ";     //syntax
		WebElement email = driver.findElement(By.xpath("//input[@type= 'email']"));
		email.sendKeys("email");

		
		String xpathWithTextMethod = "//tagname[text()='text of webelement']";   //syntax
		
		WebElement create = driver.findElement(By.xpath("//span[text() = 'Create an account']"));
		
		System.out.println(create.getText());
		
		//xpath with starts-with and contains
		
		String xpathWithStartsWith = "//tagname[starts-with(@attribute, 'partial_value')]";
		WebElement mobile = driver.findElement(By.xpath("//input[starts-with(@aria-label,'Mobile number or')]"));
		mobile.sendKeys("45687452654");
		
		String xpathWithContains = "//tagname[contains(@attribute, 'partial_value')]";
		WebElement mobile1 = driver.findElement(By.xpath("//input[contains(@aria-label,'Mobile number or')]"));
		mobile.sendKeys("45687452654");
		
		String containsWithTextMethod = "//tagname[contains(text(),'partial_value')]";
		String startsWithTextMethod = "//tagname[starts-with(text(),'partial_value')]\"";
		
		WebElement text = driver.findElement(By.xpath("//div[contains(text(),'Facebook helps you')]"));
		System.out.println(text.getText());

		//xpath with groups and indexes
		
		WebElement surname = driver.findElement(By.xpath("(//input[@class='inputtext _58mg _5dba _2ph-'])[2]"));
		surname.sendKeys("abcdef");
	}
}
	
"""

#Selenium-1stProgram
"""
from selenium import webdriver
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
# driver=webdriver.Firefox(executable_path="D:/automation/geckodriver")
# driver=webdriver.Ie(executable_path="D:/automation/IEDriverServer.exe")
driver.get("http://www.letskodeit.com")
print(driver.title)#gives title of the page
print(driver.current_url)#gives current url of the page.
print(driver.page_source)#gives HTML source code of the page.
driver.quit()#It will close all the tab, which has been opened while running this code.
#driver.close()#It closes the last tab only.
"""

#Close-&-Quit
"""
import  time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.get("http://demo.automationtesting.in/Windows.html")
driver.find_element_by_xpath("//button[text()='    click   ']").click()
time.sleep(5)
driver.close()#It closes parent browsers only.
#driver.quit()#it closes all the browsers.
"""

#NAVIGATION
"""
from selenium import webdriver
import time
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.maximize_window()
driver.get("http://demo.automationtesting.in/Windows.html")
print(driver.title)
driver.get("https://www.facebook.com/")
print(driver.title)
driver.back()
print(driver.title)
driver.forward()
print(driver.title)
driver.refresh()
time.sleep(2)
driver.quit()
"""
"""
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.maximize_window()
driver.get("http://newtours.demoaut.com/")
user_name=driver.find_element_by_name("userName")
print(user_name.is_displayed())
print(user_name.is_enabled())
user_pwd=driver.find_element_by_name("password")
print(user_name.is_displayed())
print(user_name.is_enabled())
user_name.send_keys("mercury")
user_pwd.send_keys("mercury")
driver.find_element_by_name("login").click()
button_check1=driver.find_element_by_xpath("//input[@value='roundtrip']")
print(button_check1.is_selected())
button_check2=driver.find_element_by_xpath("//input[@value='oneway']")
print(button_check2.is_selected())
"""

#IMPLICITLY WAIT
"""
from selenium import webdriver
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.maximize_window()
driver.get("http://newtours.demoaut.com/")
#Implicitly wait is applicable for all the elements of the page.
driver.implicitly_wait(10)#In seconds
print(driver.title)
assert "Welcome" in driver.title
#assert "welcome" in driver.title
user_name=driver.find_element_by_name("userName")
user_pwd=driver.find_element_by_name("password")
user_name.send_keys("mercury")
user_pwd.send_keys("mercury")
driver.find_element_by_name("login").click()
"""

#EXPLICITLY WAIT
"""
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.maximize_window()
driver.implicitly_wait(2)
driver.get("http://www.expedia.com/")
driver.find_element_by_id("package-origin-hp-package").clear()
driver.find_element_by_id("package-origin-hp-package").send_keys("NYC")
driver.find_element_by_id("package-destination-hp-package").clear()
driver.find_element_by_id("package-destination-hp-package").send_keys("SFO")
driver.find_element_by_id("package-departing-hp-package").clear()
driver.find_element_by_id("package-departing-hp-package").send_keys("04/15/2019")
driver.find_element_by_id("package-returning-hp-package").clear()
driver.find_element_by_id("package-returning-hp-package").send_keys("04/20/2019")
driver.find_element_by_id("search-button-hp-package").click()
wait=WebDriverWait(driver,10)
element=wait.until(EC.element_to_be_clickable((By.XPATH,"(//button[@class='btn-sort tab'])[1]")))
element.click()
time.sleep(5)
driver.quit()
"""

#Number of input boxes.
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.maximize_window()
driver.get("https://fs1.formsite.com/form_app/FormSite?FormId=LoadLogin&Auto")
input_boxes=driver.find_elements(By.CLASS_NAME,"auth-form__text-input")
print(len(input_boxes))
"""

#Window handles
"""
import time
from selenium import webdriver
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.maximize_window()
driver.get("https://www.facebook.com/")
print("1------------",driver.current_window_handle)
driver.find_element_by_id("cookie-use-link").click()
print("2------------",driver.current_window_handle)
driver.find_element_by_id("privacy-link").click()
print("3------------",driver.current_window_handle)
print("===============================================================================\n",driver.window_handles)
time.sleep(10)
handles=driver.window_handles
for handle in handles:
    driver.switch_to.window(handle)
    print(driver.title)
    time.sleep(5)
    if driver.title != "Facebook – log in or sign up":
        driver.close()
"""

#Scrolling down
"""
from selenium import webdriver
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.get("http://www.senojflags.com/")
#driver.execute_script("window.scrollBy(0,1000)","")
india=driver.find_element_by_xpath("(//img[contains(@alt,'India Flag')])[1]")
driver.execute_script("arguments[0].scrollIntoView;",india)
"""

#Double click
"""
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.get("http://testautomationpractice.blogspot.com/")
element=driver.find_element_by_xpath("//button[text()='Copy Text']")
action=ActionChains(driver)
action.double_click(element).perform()
time.sleep(5)
driver.quit()
"""

#Right click
"""
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
driver.get("https://swisnl.github.io/jQuery-contextMenu/demo.html")
button=driver.find_element_by_xpath("//span[text()='right click me']")
#task=driver.find_element_by_xpath("//li[contains(@class,'context-menu-item context-menu-icon context-menu-icon-edit context-menu-visible')]")
action=ActionChains(driver)
a1=action.context_click(button).perform()
time.sleep(5)
driver.quit()
"""

#YOUTUBE EXAMPLE:
#================
"""
from selenium import webdriver
import time
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe")
driver.implicitly_wait(8)
driver.get("https://www.youtube.com/watch?v=IYILCEV5j6s&list=PLUDwpEzHYYLvx6SuogA7Zhb_hZl3sln66")

print (driver.title)
time.sleep(5)
for i in range(1,10):
    locator=r"(//span[contains(@title,'Selenium with Python Tutorial {0}-')])[1]".format(i)
    print("========================================================================================================================================")
    print("locator======================",locator)
    driver.find_element_by_xpath(locator).click()
    time.sleep(8)
    print("title======================",driver.title)
    print("url======================",driver.current_url)
driver.quit()
"""

#DROP DOWN LIST:
#===============
"""
from selenium import webdriver
from selenium.webdriver.support.ui import Select
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe")
driver.get("https://www.facebook.com/")
day_parameters=Select(driver.find_element_by_id("day"))
month_parameters=Select(driver.find_element_by_id("month"))
year_parameters=Select(driver.find_element_by_id("year"))
#Select by index
day_parameters.select_by_index(5)
#Select by visible text
year_parameters.select_by_visible_text("1992")
#select by value
month_parameters.select_by_value("12")
#count all the options
print (len(day_parameters.options))
print (len(month_parameters.options))
print (len(year_parameters.options))
#capture all the options and print them
for i in day_parameters.options:
    print (i.text,end=' ')

for i in month_parameters.options:
    print (i.text,end=' ')
"""

#Handling links
#==============
"""
from selenium import webdriver
import time
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe")
driver.get("https://www.facebook.com/")
links=driver.find_elements_by_tag_name("a")
print (len(links))
for link in links:
    print (link.text)
driver.find_element_by_link_text("Cookie Policy").click()
print (driver.title)
time.sleep(3)
driver.back()
print (driver.title)
time.sleep(3)
driver.forward()
print (driver.title)
time.sleep(3)
driver.find_element_by_link_text("Cookie Policy").click()
print (driver.title)
time.sleep(3)
"""

#HANDLING FRAMES
#===============
"""
from selenium import webdriver
import time
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe")
driver.get("https://seleniumhq.github.io/selenium/docs/api/java/")
driver.maximize_window()

driver.switch_to.frame("packageListFrame")
driver.find_element_by_link_text("com.thoughtworks.selenium.webdriven").click()
time.sleep(4)
driver.switch_to.default_content()

driver.switch_to.frame("packageFrame")
driver.find_element_by_link_text("ElementFinder").click()
time.sleep(4)
driver.switch_to.default_content()

driver.switch_to.frame("classFrame")
driver.find_element_by_link_text("Instance Methods").click()
time.sleep(4)
driver.quit()
"""

#Mouse hover
#===========
"""
from selenium import webdriver
import time
from selenium.webdriver import ActionChains
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe")
# driver.get("https://www.myntra.com/")
# men=driver.find_element_by_xpath("(//a[text()='Men'])[1]")
# topwear=driver.find_element_by_xpath("(//a[text()='Topwear'])[1]")
# tshirts=driver.find_element_by_xpath("(//a[text()='T-Shirts'])[1]")
# action=ActionChains(driver)
# action.move_to_element(men).move_to_element(topwear).click().perform()
# time.sleep(10)
driver.get("https://opensource-demo.orangehrmlive.com/")
driver.find_element_by_id('txtUsername').send_keys("Admin ")
driver.find_element_by_id('txtPassword').send_keys("admin123")
driver.find_element_by_id("btnLogin").click()
admin=driver.find_element_by_xpath("//b[text()='Admin']")
administration=driver.find_element_by_id("menu_admin_UserManagement")
user=driver.find_element_by_id("menu_admin_viewSystemUsers")
action=ActionChains(driver)
action.move_to_element(admin).move_to_element(administration).move_to_element(user).click().perform()
"""

#Uploading a file
#=================
"""
from selenium import webdriver
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe")
driver.get("https://testautomationpractice.blogspot.com/")
print (driver.title)
driver.switch_to_frame(0)
driver.find_element_by_id("RESULT_FileUpload-11").send_keys("C://Users/gunjank/Desktop/CISCO_PROJECT/gunjan_KT/ltest3.txt")
"""

#Downloading a file
#==================
"""
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
c_options=Options()
c_options.add_experimental_option("prefs",{"download.default_directory": "C:\Users\gunjank\Desktop\PD"})
driver=webdriver.Chrome(executable_path="C:\Users\gunjank\Desktop\Sel_python\chromedriver.exe",chrome_options=c_options)
driver.maximize_window()
driver.get("https://www.clipconverter.cc/download/oMqyG_OC/449503778/")
driver.find_element_by_xpath("//strong[text()='Download']").click()
"""

#Reading excel file using OpenPyXl
#=================================
"""
import openpyxl
workbook=openpyxl.load_workbook("C:\Users\gunjank\Desktop\PD\data.xlsx")
sheet=workbook.active     #workbook.get_sheet_by_name("sheet1")
rows=sheet.max_row
columns=sheet.max_column
print ("rows===",rows)
print ("columns===",columns)
for r in range(1, rows+1):
    for c in  range(1,columns+1):
        print(sheet.cell(row=r, column=c).value)
    print " "
"""

"""
import openpyxl
path="C:\Users\gunjank\Desktop\PD\data.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active     #workbook.get_sheet_by_name("sheet1")
for r in range(1, 4):
    for c in  range(1,6):
        sheet.cell(row=r, column=c).value="Welcome"
workbook.save(path)
"""

#Keywords
#=========
"""
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
driver=webdriver.Chrome(executable_path="D:/automation/chromedriver")
# driver.get("https://opensource-demo.orangehrmlive.com/")
# driver.find_element_by_id('txtUsername').send_keys("Admin ")
driver.get("https://www.myntra.com/")
time.sleep(3)
actions=ActionChains(driver)
actions.send_keys(Keys.TAB).perform()
time.sleep(1)
actions.send_keys(Keys.TAB).perform()
time.sleep(1)
actions.send_keys(Keys.TAB).send_keys(Keys.ENTER).perform()
time.sleep(3)
driver.quit()
"""