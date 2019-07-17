import openpyxl
import logging

logging.basicConfig(filename="test.log",
                    format='%(asctime)s: %(levelname)s: %(message)s',
                    datefmt='%d/%m/%y %I/%M/%S %p'
                    )
logger=logging.getLogger()
logger.setLevel(logging.INFO)


def getRowsCount(data_file,sheet_name):
    workbook=openpyxl.load_workbook(data_file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    logger.info("counting rows")
    return sheet.max_row

def getColumnsCount(data_file,sheet_name):
    workbook=openpyxl.load_workbook(data_file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    logger.info("counting columns")
    return sheet.max_column

def readData(data_file,sheet_name,row_num,column_num):
    logger.info("reading data")
    workbook=openpyxl.load_workbook(data_file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num,column=column_num).value

def writeData(data_file,sheet_name,row_num,column_num,output_data):
    logger.info("writing data")
    workbook=openpyxl.load_workbook(data_file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num,column=column_num).value=output_data
    workbook.save(data_file)
